import os
import pandas as pd
import openpyxl

def mask_email(email):
    if pd.isna(email) or "@" not in email:
        return email
    name, domain = email.split("@", 1)
    return f"{name[0]}{'*' * (len(name)-1)}@{domain}"

def mask_phone(phone):
    if pd.isna(phone) or not isinstance(phone, str):
        return phone
    return phone[:-2] + "**" if len(phone) > 2 else "**"

def mask_social(url):
    if pd.isna(url) or not isinstance(url, str):
        return url
    return url.split("/")[-1][:2] + "****"

def mask_dataframe(df):
    for col in df.columns:
        if "email" in col.lower():
            df[col] = df[col].apply(mask_email)
        elif "phone" in col.lower() or "tel" in col.lower():
            df[col] = df[col].apply(mask_phone)
        elif any(s in col.lower() for s in ["facebook", "instagram", "linkedin", "x", "twitter"]):
            df[col] = df[col].apply(mask_social)
    return df

def process_csv(file_path, output_path):
    df = pd.read_csv(file_path)
    df = mask_dataframe(df)
    df.to_csv(output_path, index=False)

def process_xlsx(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    val = cell.value.lower()
                    if "@" in val:
                        cell.value = mask_email(cell.value)
                    elif any(s in val for s in ["facebook", "instagram", "linkedin", "x.com", "twitter"]):
                        cell.value = mask_social(cell.value)
                    elif any(char.isdigit() for char in val) and len(val) >= 7:
                        cell.value = mask_phone(cell.value)
    wb.save(output_path)

def main():
    input_folder = "demo_inputs"
    output_folder = "demo_outputs"

    if not os.path.exists(input_folder):
        print(f"❌ La carpeta '{input_folder}' no existe. Por favor, créala y añade archivos CSV o XLSX.")
        return

    os.makedirs(output_folder, exist_ok=True)

    for filename in os.listdir(input_folder):
        input_path = os.path.join(input_folder, filename)
        output_path = os.path.join(output_folder, filename)

        if filename.endswith(".csv"):
            process_csv(input_path, output_path)
            print(f"✅ Procesado CSV: {filename}")
        elif filename.endswith(".xlsx"):
            process_xlsx(input_path, output_path)
            print(f"✅ Procesado Excel: {filename}")

if __name__ == "__main__":
    main()
