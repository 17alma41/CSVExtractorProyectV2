import os
import pandas as pd
from openpyxl import load_workbook

# Ruta base del servidor y ruta de salida
BASE_PATH = r"\\SERVIDOR3001\Central\OMK\Publicar"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "data", "outputs")
OUTPUT_FILE = os.path.join(OUTPUT_PATH, "Resumen_Publicar.xlsx")

# Crear carpeta de salida si no existe
os.makedirs(OUTPUT_PATH, exist_ok=True)

# Columnas del archivo de salida
columns = [
    "NombreArchivo", "País", "Sector", "NRegistros", "NMail", "NTelefonos", "Nrrss",
    "URLDescarga", "URLDemo", "Foto1", "Captura1", "Captura2", "Captura3"
]
resultados = []

def extraer_métricas(path_excel):
    try:
        wb = load_workbook(path_excel, data_only=True)
        if "statistics" not in wb.sheetnames:
            print(f"⚠️ La hoja 'statistics' no está en {path_excel}")
            return 0, 0, 0, 0
        ws = wb["statistics"]

        headers = [cell.value for cell in ws[1]]
        values = [cell.value for cell in ws[2]]

        metrics = dict(zip(headers, values))

        n_registros = int(metrics.get("Number of companies", 0) or 0)
        n_mail = int(metrics.get("Number of emails (unique)", 0) or 0)
        n_telefonos = int(metrics.get("Mobile phones", 0) or 0)
        n_rrss = int(metrics.get("Number of social networks", 0) or 0)

        return n_registros, n_mail, n_telefonos, n_rrss

    except Exception as e:
        print(f"❌ Error leyendo {path_excel}: {e}")
        return 0, 0, 0, 0


# Recorre carpetas por país
for pais in os.listdir(BASE_PATH):
    pais_path = os.path.join(BASE_PATH, pais)
    if not os.path.isdir(pais_path):
        continue

    archivos = os.listdir(pais_path)
    excel_files = [f for f in archivos if f.lower().endswith(".xlsx")]
    jpg_files = [f for f in archivos if f.lower().endswith(".jpg")]

    for excel in excel_files:
        ruta_excel = os.path.join(pais_path, excel)
        nombre_archivo = os.path.splitext(excel)[0]
        sector = nombre_archivo.split("-")[1] if "-" in nombre_archivo else ""

        print(f"📄 Procesando: {excel} en {pais}")

        n_reg, n_mail, n_tel, n_rrss = extraer_métricas(ruta_excel)

        capturas = jpg_files[:3] + [""] * (3 - len(jpg_files))

        fila = [
            nombre_archivo,
            pais,
            sector.replace("_", " "),
            n_reg,
            n_mail,
            n_tel,
            n_rrss,
            "", "", "",  # URLDescarga, URLDemo, Foto1
            *capturas
        ]
        resultados.append(fila)

# Guardar en Excel
df = pd.DataFrame(resultados, columns=columns)
df.to_excel(OUTPUT_FILE, index=False)
print(f"\n✅ Archivo generado correctamente en:\n{OUTPUT_FILE}")
