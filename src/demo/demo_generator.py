"""
demo_generator.py - Módulo para enmascarar datos sensibles y generar archivos demo.
"""

import os
import pandas as pd
import openpyxl
from pathlib import Path
from src.settings import DEMO_INPUTS_DIR, DEMO_OUTPUTS_DIR
from src.utils.status_manager import load_status, update_status, is_stage_done, log_error

def mask_email(email):
    if pd.isna(email) or "@" not in str(email):
        return email
    name, domain = str(email).split("@", 1)
    if not name:
        return email
    return f"{name[0]}{'*' * (len(name)-1)}@{domain}"

def mask_phone(phone):
    if pd.isna(phone) or not isinstance(phone, str):
        return phone
    return phone[:-2] + "**" if len(phone) > 2 else "**"

def mask_social(url):
    if pd.isna(url) or not isinstance(url, str):
        return url
    return url.split("/")[-1][:2] + "****"

def mask_dataframe(df):
    for col in df.columns:
        if "email" in col.lower():
            df[col] = df[col].apply(mask_email)
        elif "phone" in col.lower() or "tel" in col.lower():
            df[col] = df[col].apply(mask_phone)
        elif any(s in col.lower() for s in ["facebook", "instagram", "linkedin", "x", "twitter"]):
            df[col] = df[col].apply(mask_social)
    return df

def process_csv(file_path, output_path):
    df = pd.read_csv(file_path)
    df = mask_dataframe(df)
    df.to_csv(output_path, index=False)

def process_xlsx(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    val = cell.value.lower()
                    if "@" in val:
                        cell.value = mask_email(cell.value)
                    elif any(s in val for s in ["facebook", "instagram", "linkedin", "x.com", "twitter"]):
                        cell.value = mask_social(cell.value)
                    elif any(char.isdigit() for char in val) and len(val) >= 7:
                        cell.value = mask_phone(cell.value)
    wb.save(output_path)

def run_demo(overwrite=False, test_mode=False, resume=False):
    """
    Genera versiones demo enmascaradas de los archivos en DEMO_INPUTS_DIR.
    """
    input_folder = DEMO_INPUTS_DIR
    output_folder = DEMO_OUTPUTS_DIR
    if not input_folder.exists():
        print(f"❌ La carpeta '{input_folder}' no existe. Por favor, créala y añade archivos CSV o XLSX.")
        return
    output_folder.mkdir(parents=True, exist_ok=True)
    for filename in os.listdir(input_folder):
        try:
            if not overwrite and is_stage_done(filename, 'demo_generated'):
                print(f"[SKIP] {filename} ya enmascarado.")
                continue
            src = input_folder / filename
            if not filename.lower().endswith((".csv", ".xlsx")):
                continue
            stem = src.stem
            ext  = src.suffix
            dest_name = f"{stem}_demo{ext}"
            dest = output_folder / dest_name
            if ext == ".csv":
                process_csv(src, dest)
                print(f"✅ Procesado CSV:  {filename} → {dest_name}")
            else:
                process_xlsx(src, dest)
                print(f"✅ Procesado Excel: {filename} → {dest_name}")
            update_status(filename, 'demo_generated', True)
        except Exception as e:
            log_error(filename, 'demo', '', str(e))
